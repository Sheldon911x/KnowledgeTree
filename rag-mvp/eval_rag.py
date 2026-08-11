# -*- coding: utf-8 -*-
"""
RAG 评测脚本（检索召回层）
========================
目的：把斯达高问答数据集本身当作「黄金集」——每条记录自带问题和答案，
      用其「问题」当查询，检验 RAG 能否把对应的那条问答块检索到 top-K 里。

指标：Recall@K（召回率） = 命中的查询数 / 总查询数
      命中定义：检索 top-K 中，存在来源标签以 `#<该记录编号>` 结尾的块。

用法：
  python eval_rag.py                                   # 默认评测 kbs/sidaogao.json，对比 dense/bm25/hybrid
  python eval_rag.py --k 5                             # 评测 Recall@5
  python eval_rag.py --modes dense,hybrid              # 只对比指定模式
  python eval_rag.py --store kbs/sidaogao.json --xlsx "D:\飞书文件下载\改版-斯达高知识库问答数据集.xlsx"

说明：
  - 默认用「原问题」作查询，召回通常会偏高（块内就含原问题），这是基线。
  - 更有价值的评测是用「改写 / paraphrase」的问题测泛化：可把改写问题+应命中编号
    做成 golden.csv（两列：问题,应命中编号），再用 --golden 指定。
  - 评测只跑检索层（不调生成模型），所以很快；生成质量请结合 ask --debug 人工看。
"""

import os
import sys
import json
import argparse

import numpy as np

# 复用 rag_mvp 的检索 / 嵌入逻辑，避免在两份代码里各写一遍
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import rag_mvp as R

DEFAULT_STORE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kbs", "sidaogao.json")
DEFAULT_XLSX  = r"D:\飞书文件下载\改版-斯达高知识库问答数据集.xlsx"


def load_golden_from_xlsx(xlsx_path):
    """返回 [(rid, question), ...]，直接从问答数据集抽取。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # 跳过表头
        for r in rows:
            if not r:
                continue
            rid = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
            q   = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            if not rid or not q:
                continue
            out.append((rid, q))
    return out


def load_golden_from_csv(csv_path):
    """golden.csv：两列 问题,应命中编号（首行可为表头，自动跳过）。用于改写问题评测。"""
    import csv
    out = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        started = False
        for row in reader:
            if len(row) < 2:
                continue
            q, rid = row[0].strip(), row[1].strip()
            if not started:
                started = True
                # 首行若是表头则跳过
                if q in ("问题", "question", "query", "q") or rid in ("应命中编号", "编号", "id", "rid"):
                    continue
            if q and rid:
                out.append((rid, q))
    return out


def main():
    ap = argparse.ArgumentParser(description="RAG 检索召回评测")
    ap.add_argument("--store", default=DEFAULT_STORE, help="索引文件路径")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="问答数据集 xlsx（默认黄金集来源）")
    ap.add_argument("--golden", default=None, help="可选的 golden.csv（改写问题评测）")
    ap.add_argument("--k", type=int, default=R.TOP_K, help="Recall@K 的 K")
    ap.add_argument("--modes", default="dense,bm25,hybrid", help="对比的检索模式，逗号分隔")
    args = ap.parse_args()

    if not os.path.exists(args.store):
        print(f"索引不存在：{args.store}，请先 build --store 该路径"); sys.exit(1)

    store = R._load_store(args.store)

    if args.golden:
        golden = load_golden_from_csv(args.golden)
        src = args.golden
    else:
        if not os.path.exists(args.xlsx):
            print(f"黄金集文件不存在：{args.xlsx}"); sys.exit(1)
        golden = load_golden_from_xlsx(args.xlsx)
        src = args.xlsx

    print(f"索引块数：{len(store)}  黄金查询数：{len(golden)}  K={args.k}")
    print(f"黄金集来源：{src}\n")

    modes = [m.strip().lower() for m in args.modes.split(",") if m.strip()]

    # 一次性批量嵌入所有查询问题（避免逐条请求 Ollama 超时）
    print("批量嵌入查询问题…")
    queries = [q for _, q in golden]
    q_vecs = R.embed_batch(queries)
    print("嵌入完成，开始评测…\n")

    results = {}
    misses = {m: [] for m in modes}
    for mode in modes:
        hit = 0
        for i, (rid, _) in enumerate(golden):
            top = R.retrieve(queries[i], store, k=args.k, mode=mode, q_vec=q_vecs[i])
            hit_rids = [t["source"].split("#")[-1] for t in top]
            if rid in hit_rids:
                hit += 1
            else:
                misses[mode].append((rid, hit_rids))
        results[mode] = hit / len(golden)
        print(f"  [{mode:7}] Recall@{args.k} = {results[mode]:.3f}  ({hit}/{len(golden)})")

    print("\n" + "=" * 52)
    print("三种检索模式召回对比：")
    best = max(results, key=results.get)
    for m in modes:
        mark = "  <-- 最佳" if m == best else ""
        print(f"  {m:7} : {results[m]:.3f}{mark}")
    print("=" * 52)

    # 列出 hybrid（或首个模式）的未命中样本，便于调参定位
    show_mode = "hybrid" if "hybrid" in modes else modes[0]
    if misses[show_mode]:
        print(f"\n[{show_mode}] 未命中样本（前 15 条）：问题编号 -> 实际命中的来源编号")
        for rid, got in misses[show_mode][:15]:
            print(f"  #{rid:>4}  ->  命中 {got}")
        print(f"  ... 共 {len(misses[show_mode])} 条未命中")
        print("提示：未命中通常是该问题表述特殊 / 答案块过长被稀释；可调 CHUNK、TOP_K、HYBRID_ALPHA 后重测对比。")
    else:
        print(f"\n[{show_mode}] 全部命中，Recall@{args.k}=1.000")


if __name__ == "__main__":
    main()
