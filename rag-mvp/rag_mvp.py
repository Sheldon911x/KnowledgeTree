# -*- coding: utf-8 -*-
"""
本地 RAG MVP（零成本 / 纯本地 · 含 BM25 混合检索与调试模式）
==============================================================
一条命令建索引，一条命令问答。所有计算都在本机完成：
  - 嵌入模型：bge-m3       （文本 -> 向量，已 pull）
  - 生成模型：qwen2.5:3b    （基于检索内容作答，已 pull）
  - 向量库：numpy 余弦相似度，落盘 JSON，无外部服务
  - 检索：支持 dense / bm25 / 混合(hybrid) 三种模式

RAG 五个环节（对应五个函数）：
  1) load_document  加载
  2) chunk_text     切分
  3) embed          嵌入
  4) retrieve       检索（dense / bm25 / 混合）
  5) generate_answer 生成

用法：
  python rag_mvp.py build                 # 建索引（默认 DOCS_DIR -> vector_store.json，只需一次）
  python rag_mvp.py build --docs 文件或目录 --store 索引.json   # 单独建一个知识库
  python rag_mvp.py ask "问题"            # 单次问答（默认库）
  python rag_mvp.py ask "问题" --store 索引.json              # 对指定库问答
  python rag_mvp.py ask "问题" --debug    # 单次问答 + 显示各路打分（调试用）
  python rag_mvp.py chat  [--store 索引.json]   # 交互式问答
  python rag_mvp.py inspect [--store 索引.json]  # 打印索引统计与当前配置（调试入口）

提示：用 --docs / --store 可以把不同数据源建成互相独立的索引库，互不干扰。
"""

import os
import sys
import re
import json
import math
import glob
import argparse
from collections import Counter
from html.parser import HTMLParser

import numpy as np
import requests

# ========================= 配置区（所有开关都在这里） =========================
OLLAMA_BASE = "http://localhost:11434"   # Ollama 桌面版默认地址
EMBED_MODEL  = "bge-m3"                   # 嵌入模型（你已 pull）
GEN_MODEL    = "qwen2.5:3b"               # 生成模型（你已 pull）

DOCS_DIR     = r"D:\rag_test_kb"          # 测试文档目录
STORE_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vector_store.json")

# —— 切分（chunk）设置 ——
CHUNK_SIZE     = 600      # 每块字符数
CHUNK_OVERLAP  = 100      # 块与块重叠字符数（避免切断语义）

# —— 检索设置 ——
TOP_K          = 4        # 检索返回最相关的几块
RETRIEVAL_MODE = "hybrid" # "dense"(纯向量) | "bm25"(纯关键词) | "hybrid"(混合，推荐)
HYBRID_ALPHA   = 0.6      # 混合时稠密向量权重；BM25 权重 = 1 - ALPHA
# ==============================================================================


# ---------- 1) 加载：文件 -> 纯文本 ----------
class _HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []
    def handle_data(self, data):
        self.parts.append(data)

def _strip_html(html_text: str) -> str:
    p = _HTMLStripper()
    p.feed(html_text)
    return "\n".join(p.parts)

def load_document(path: str) -> str:
    """按扩展名读取；md/txt/json/csv 直接读，html 去标签。pdf/docx 见 requirements.txt。"""
    ext = os.path.splitext(path)[1].lower()
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        raw = f.read()
    if ext == ".html":
        return _strip_html(raw)
    return raw

def _read_xlsx_rows(path: str) -> list:
    """Excel 问答数据集：一行 = 一条记录（编号/问题/答案）。按需 pip install openpyxl。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    units = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # 跳过表头
        for r in rows:
            if not r:
                continue
            rid = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
            q   = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            a   = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
            if not q and not a:
                continue
            text = f"问题：{q}\n答案：{a}"
            label = f"{os.path.basename(path)}#{rid}" if rid else os.path.basename(path)
            units.append((text, label))
    return units

def iter_units(path: str) -> list:
    """返回 [(文本, 来源标签), ...]。
    - .xlsx 问答数据集：一行一条（不按字符硬切，避免切断问答）
    - 其余文件：整体读入后按 chunk 切分"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            return _read_xlsx_rows(path)
        except Exception as e:
            print(f"  [跳过] xlsx 读取失败 {os.path.basename(path)}: {e}")
            return []
    return [(c, os.path.basename(path)) for c in chunk_text(load_document(path))]

# ---------- 2) 切分：长文本 -> 有重叠的小块 ----------
def chunk_text(text: str, size=CHUNK_SIZE, overlap=CHUNK_OVERLAP) -> list:
    """滑动窗口切分；短文本（不足 size）整体作为一块。"""
    text = text.strip()
    if not text:
        return []
    if len(text) <= size:
        return [text]
    step = max(1, size - overlap)
    return [c for c in (text[i:i + size] for i in range(0, len(text), step)) if c.strip()]

# ---------- 3) 嵌入：文本 -> 向量 ----------
def _normalize(vec: np.ndarray) -> np.ndarray:
    norm = np.linalg.norm(vec)
    return vec / norm if norm > 0 else vec

def embed(text: str) -> np.ndarray:
    """单条嵌入（用于检索时的问题向量）。"""
    resp = requests.post(
        f"{OLLAMA_BASE}/api/embeddings",
        json={"model": EMBED_MODEL, "prompt": text},
        timeout=60,
    )
    resp.raise_for_status()
    return _normalize(np.array(resp.json()["embedding"], dtype=np.float32))

def embed_batch(texts: list, batch_size: int = 32) -> list:
    """批量嵌入（建索引用，大幅减少 HTTP 往返，速度远快于逐条调用）。"""
    out = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        resp = requests.post(
            f"{OLLAMA_BASE}/api/embed",
            json={"model": EMBED_MODEL, "input": batch},
            timeout=300,
        )
        resp.raise_for_status()
        for vec in resp.json()["embeddings"]:
            out.append(_normalize(np.array(vec, dtype=np.float32)))
    return out

# ---------- BM25（关键词检索，纯 Python 实现，无需额外依赖） ----------
# 中文按字、英文/数字按词切分，作为 BM25 的词条
_TOKEN_RE = re.compile(r"[a-zA-Z0-9_]+|[\u4e00-\u9fff]")
def tokenize(text: str) -> list:
    return _TOKEN_RE.findall(text.lower())

class BM25:
    def __init__(self, corpus_tokens, k1=1.5, b=0.75):
        self.k1, self.b = k1, b
        self.docs = corpus_tokens
        self.N = len(corpus_tokens)
        self.df = {}
        for doc in corpus_tokens:
            for t in set(doc):
                self.df[t] = self.df.get(t, 0) + 1
        self.avgdl = sum(len(d) for d in corpus_tokens) / max(1, self.N)
        self.idf = {t: math.log((self.N - df + 0.5) / (df + 0.5) + 1)
                    for t, df in self.df.items()}
    def scores(self, query_tokens):
        out = []
        for doc in self.docs:
            score, dl, freq = 0.0, len(doc), Counter(doc)
            for t in query_tokens:
                if t in self.idf and freq.get(t, 0):
                    f = freq[t]
                    score += self.idf[t] * (f * (self.k1 + 1)) / \
                             (f + self.k1 * (1 - self.b + self.b * dl / self.avgdl))
            out.append(score)
        return out

# ---------- 构建索引 ----------
def build_index(docs_dir=None, store_path=None):
    docs_dir = docs_dir or DOCS_DIR
    store_path = store_path or STORE_PATH
    # --docs 可以是目录（扫描全部文件）或单个文件（如某个 xlsx）
    if os.path.isdir(docs_dir):
        files = sorted(glob.glob(os.path.join(docs_dir, "*")))
        files = [f for f in files if os.path.isfile(f)]
    elif os.path.isfile(docs_dir):
        files = [docs_dir]
    else:
        files = []
    if not files:
        print(f"在 {docs_dir} 没找到任何文件，请检查 --docs 路径。")
        return
    # 先收集所有 (文本, 来源标签)，再批量嵌入（避免逐条调用过慢）
    all_texts, all_labels, nfiles = [], [], 0
    for f in files:
        units = iter_units(f)
        if not units:
            continue
        nfiles += 1
        for text, label in units:
            all_texts.append(text)
            all_labels.append(label)
        print(f"  [扫描] {os.path.basename(f)} -> {len(units)} 块")
    print(f"共 {len(all_texts)} 块，开始批量嵌入（CPU 上可能需几分钟）…")
    vecs = embed_batch(all_texts)
    store = [{"source": lbl, "chunk_id": 0, "text": txt, "vector": vec.tolist()}
              for txt, lbl, vec in zip(all_texts, all_labels, vecs)]
    with open(store_path, "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False)
    print(f"\n索引构建完成：{nfiles} 文件，{len(store)} 块 -> {store_path}")

# ---------- 4) 检索：dense / bm25 / hybrid ----------
def _load_store(store_path=None):
    store_path = store_path or STORE_PATH
    if not os.path.exists(store_path):
        print(f"索引不存在（{store_path}），请先运行：python rag_mvp.py build --store {store_path}")
        sys.exit(1)
    with open(store_path, "r", encoding="utf-8") as f:
        return json.load(f)

def _minmax(a: np.ndarray) -> np.ndarray:
    if a.max() == a.min():
        return np.zeros_like(a)
    return (a - a.min()) / (a.max() - a.min())

def retrieve(query: str, store: list, k=TOP_K, debug=False, mode=None, alpha=HYBRID_ALPHA, q_vec=None) -> list:
    mode = (mode or RETRIEVAL_MODE).lower()
    # 稠密（向量）打分
    if q_vec is None:
        q_vec = embed(query)
    dense = np.array([it["vector"] for it in store], dtype=np.float32) @ q_vec
    # 稀疏（BM25）打分
    bm25 = BM25([tokenize(it["text"]) for it in store])
    bm25_scores = np.array(bm25.scores(tokenize(query)), dtype=np.float32)
    # 各自归一化后融合
    dense_n, bm25_n = _minmax(dense), _minmax(bm25_scores)
    if mode == "dense":
        combined = dense_n
    elif mode == "bm25":
        combined = bm25_n
    else:
        combined = alpha * dense_n + (1 - alpha) * bm25_n
    top_idx = np.argsort(-combined)[:k]
    results = []
    for idx in top_idx:
        item = dict(store[idx])
        item["score"] = float(combined[idx])
        if debug:
            item["dense"] = float(dense[idx])
            item["bm25"] = float(bm25_scores[idx])
        results.append(item)
    return results

# ---------- 5) 生成 ----------
def generate_answer(query: str, contexts: list) -> str:
    context_block = "\n\n".join(
        f"[来源 {c['source']} | 相似度 {c['score']:.3f}]\n{c['text']}" for c in contexts
    )
    prompt = (
        "你是一个严格基于【知识库】内容回答问题的助手。\n"
        "规则：\n"
        "1) 只使用知识库里给出的信息作答；\n"
        "2) 如果知识库中没有相关信息，明确回答“知识库中没有相关信息”，不要编造；\n"
        "3) 回答使用简体中文，简洁准确，可引用来源文件名。\n\n"
        f"【知识库】\n{context_block}\n\n"
        f"【问题】\n{query}\n\n"
        "【回答】"
    )
    resp = requests.post(
        f"{OLLAMA_BASE}/api/generate",
        json={"model": GEN_MODEL, "prompt": prompt, "stream": False},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.json()["response"].strip()

# ---------- 调试：索引统计 + 当前配置 ----------
def cmd_inspect(store_path=None):
    store = _load_store(store_path)
    lengths = [len(it["text"]) for it in store]
    sources = sorted({it["source"].split("#")[0] for it in store})
    print("=" * 60)
    print("【当前配置】")
    for k, v in [("OLLAMA_BASE", OLLAMA_BASE), ("EMBED_MODEL", EMBED_MODEL),
                 ("GEN_MODEL", GEN_MODEL), ("DOCS_DIR", DOCS_DIR),
                 ("STORE_PATH", store_path or STORE_PATH),
                 ("CHUNK_SIZE", CHUNK_SIZE), ("CHUNK_OVERLAP", CHUNK_OVERLAP),
                 ("TOP_K", TOP_K), ("RETRIEVAL_MODE", RETRIEVAL_MODE),
                 ("HYBRID_ALPHA", HYBRID_ALPHA)]:
        print(f"  {k:15} = {v}")
    print("-" * 60)
    print("【索引统计】")
    print(f"  文件数      : {len(sources)}")
    print(f"  文本块数    : {len(store)}")
    print(f"  块长度 平均 : {sum(lengths)//len(lengths)} 字符")
    print(f"  块长度 最小 : {min(lengths)} / 最大 : {max(lengths)} 字符")
    print("=" * 60)
    print("调试建议：")
    print("  - 看检索效果：python rag_mvp.py ask \"问题\" --debug")
    print("  - 换检索模式：把 RETRIEVAL_MODE 改成 dense / bm25 / hybrid")
    print("  - 调混合权重：改 HYBRID_ALPHA（越大越偏语义，越小越偏关键词）")
    print("  - 调切块    ：改 CHUNK_SIZE / CHUNK_OVERLAP 后重新 build")

# ---------- 交互入口 ----------
def cmd_ask(query: str, debug=False, store_path=None):
    store = _load_store(store_path)
    ctx = retrieve(query, store, debug=debug)
    ans = generate_answer(query, ctx)
    print("\n" + "=" * 60)
    print("检索到的参考片段：")
    for c in ctx:
        line = f"  - {c['source']} (综合 {c['score']:.3f})"
        if debug:
            line += f"  [dense {c.get('dense',0):.3f} | bm25 {c.get('bm25',0):.3f}]"
        print(line)
    print("=" * 60)
    print("\n回答：\n" + ans)

def cmd_chat(store_path=None):
    store = _load_store(store_path)
    print("进入问答模式（输入 exit / quit 退出）\n")
    while True:
        try:
            q = input("你：").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n再见。"); break
        if not q:
            continue
        if q.lower() in ("exit", "quit", "退出"):
            print("再见。"); break
        ctx = retrieve(q, store)
        print("\n回答：\n" + generate_answer(q, ctx))
        print("-" * 60)

def main():
    parser = argparse.ArgumentParser(description="本地 RAG MVP")
    sub = parser.add_subparsers(dest="cmd")
    p_build = sub.add_parser("build", help="构建向量索引（只需一次）")
    p_build.add_argument("--docs", default=None, help="要建库的文件或目录（默认 DOCS_DIR）")
    p_build.add_argument("--store", default=None, help="索引保存路径（默认 vector_store.json）")
    p_ask = sub.add_parser("ask", help="单次问答")
    p_ask.add_argument("query", help="你的问题")
    p_ask.add_argument("--debug", action="store_true", help="显示各路检索打分")
    p_ask.add_argument("--store", default=None, help="使用指定的索引文件")
    p_chat = sub.add_parser("chat", help="交互式问答")
    p_chat.add_argument("--store", default=None, help="使用指定的索引文件")
    p_inspect = sub.add_parser("inspect", help="打印索引统计与配置（调试入口）")
    p_inspect.add_argument("--store", default=None, help="查看指定索引的统计")
    args = parser.parse_args()
    if args.cmd == "build":
        build_index(args.docs, args.store)
    elif args.cmd == "ask":
        cmd_ask(args.query, debug=args.debug, store_path=args.store)
    elif args.cmd == "chat":
        cmd_chat(args.store)
    elif args.cmd == "inspect":
        cmd_inspect(args.store)
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
